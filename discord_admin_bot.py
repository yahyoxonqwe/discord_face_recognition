import re
import os
import sqlite3
import pandas as pd
from io import BytesIO
from datetime import datetime , timedelta , time
from config import *
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from xlsxwriter.utility import xl_col_to_name
from openpyxl.utils.dataframe import dataframe_to_rows
import discord
from discord.ext import commands
from discord import File

admin_id = [897019961460019220 , 1344616926130864179]


# ---------- Discord Bot ----------
intents = discord.Intents.default()
intents.messages = True
intents.dm_messages = True
bot = commands.Bot(command_prefix='!', intents=intents)


# SQLite database connection
db_path = 'xb_database.db'

pattern_single_day = re.compile(r'^[0-9]{2}$')
pattern_full_date = re.compile(r'^[0-9]{2}-\d{2}-\d{4}$')
pattern_day_range = re.compile(r'^[0-9]{2}-[0-9]{2}$')
pattern_date_range = re.compile(r'^\b\d{2}-\d{2}-\d{4}-\b\d{2}-\d{2}-\d{4}$')


green = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
orange = openpyxl.styles.PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')


def query_data_to_dataframe(target_date):
    conn = sqlite3.connect(db_path)
    start = datetime.combine(target_date, time(4,0,15))

    query = "SELECT name,arrival_time,departure_time,image FROM users WHERE arrival_time BETWEEN ? AND ? AND name NOT LIKE 'unknown%'"
    query2 = "SELECT name,arrival_time,departure_time,image FROM users WHERE arrival_time BETWEEN ? AND ? AND name LIKE 'unknown%'"

    df1 = pd.read_sql_query(query, conn, params=(start, start + timedelta(days=1)))
    df2 = pd.read_sql_query(query2, conn, params=(start, start + timedelta(days=1)))
    return df1,df2

#     10 minutgacha kesh qolganlardan -0.05
#     1 soatgacha -0,1
#     2 soatgacha -0,2
#     3 soatcgacha -0,3
#     4 soat -0,4
#     4 soatdan kattaga -0,5 olinib tashlanadi.

#     1. Shanba kunini borderlarini QALIN qilib yuklaydigan bo'lay
#     2. Har kunni yoniga bitta column qo'shib ber. Bu uning ayriladigan ballari bo'lsin. Kelmasa -0,5, Kesh qolsa tepadagilar, Kesh kelib, erta ketsa KELISH-KETISH vaqtini 9 soatdan farxiga qarab tepadagi jadval bo'yicha ballarni ayirasan, erta kelib erta ketsa ham xuddi shu sistemani qilasan.
#     3. Hamma kunlarni 9 soat farxi qilib qo'y. Juda erta kelib 9 soat ishlasa to'liq ishlagan bo'lsin.
#     Keyin xohlagan kuni 10:30dan oldin kelib 9 soat (1soat obed) ishlagan odam 18:00dan oldin qaytsa ham bo'ladi. Masalan 7:30da kelib 16:30da qaytsa bo'ladi.




def get_ball(start, arrival_time, departure_time):
    
    limit = timedelta(hours=8, minutes=59)
    
    # agar 10:30 dal oldin kelgan bo'lsa
    if arrival_time < start + timedelta(hours=6, minutes=30):
        if departure_time - arrival_time > limit or departure_time > start + timedelta(hours=13, minutes=59):
            return "0"
        elif departure_time - arrival_time > limit - timedelta(minutes=10):
            return "-0.05"
        elif departure_time - arrival_time > limit - timedelta(hours=1):
            return "-0.1"
        elif departure_time - arrival_time > limit - timedelta(hours=2):
            return "-0.2"
        elif departure_time - arrival_time > limit - timedelta(hours=3):
            return "-0.3"
        elif departure_time - arrival_time > limit - timedelta(hours=4):
            return "-0.4"
        else:
            return "-0.5"
    elif departure_time < start + timedelta(hours=13, minutes=59):
#         limit = timedelta(hours=7, minutes=30)
        if departure_time - arrival_time > limit - timedelta(minutes=10):
            return "-0.05"
        elif departure_time - arrival_time > limit - timedelta(hours=1):
            return "-0.1"
        elif departure_time - arrival_time > limit - timedelta(hours=2):
            return "-0.2"
        elif departure_time - arrival_time > limit - timedelta(hours=3):
            return "-0.3"
        elif departure_time - arrival_time > limit - timedelta(hours=4):
            return "-0.4"
        else:
            return "-0.5"
    else:
        if arrival_time < start + timedelta(hours=6, minutes = 40):
            return "-0.05"
        elif arrival_time < start + timedelta(hours=7, minutes = 30):
            return "-0.1"
        elif arrival_time < start + timedelta(hours=8, minutes = 30):
            return "-0.2"        
        elif arrival_time < start + timedelta(hours=9, minutes = 30):
            return "-0.3"        
        elif arrival_time < start + timedelta(hours=10, minutes = 30):
            return "-0.4"        
        else:
            return "-0.5"
    

@bot.command(name='start')
async def start(ctx):
    # This will send the prompt in the same channel
    await ctx.send("DD | DD-DD | DD-MM-YYYY | DD-MM-YYYY-DD-MM-YYYY format kiriting.")
PROMPT = "DD | DD-DD | DD-MM-YYYY | DD-MM-YYYY-DD-MM-YYYY format kiriting."
@bot.event
async def on_message(message):
    # ignore bots (including itself)
    if message.author.bot:
        return

    # only allow admins
    if message.author.id not in admin_id:
        await message.reply("Uzur , siz admin emassiz ")

    text = message.content.strip()
    days = []

    # 1) Single day of current month: "DD"
    if pattern_single_day.match(text):
        try:
            days.append(datetime(
                year=datetime.now().year,
                month=datetime.now().month,
                day=int(text)
            ))
        except ValueError:
            await message.reply(PROMPT)
            return

    # 2) Full date: "DD-MM-YYYY"
    elif pattern_full_date.match(text):
        try:
            days.append(datetime.strptime(text, "%d-%m-%Y"))
        except ValueError:
            await message.reply(PROMPT)
            return

    # 3) Day-to-day range within current month: "DD-DD"
    elif pattern_day_range.match(text):
        start_d, end_d = map(int, text.split('-'))
        try:
            start = datetime(datetime.now().year, datetime.now().month, start_d)
            end   = datetime(datetime.now().year, datetime.now().month, end_d)
        except ValueError:
            await message.reply(PROMPT)
            return

        if end <= start:
            await message.reply(PROMPT)
            return

        days = [start + timedelta(days=i) for i in range((end - start).days + 1)]

    # 4) Full date-to-date range: "DD-MM-YYYY-DD-MM-YYYY"
    elif pattern_date_range.match(text):
        parts = text.split('-')
        start_str = '-'.join(parts[:3])
        end_str   = '-'.join(parts[3:])
        try:
            start = datetime.strptime(start_str, "%d-%m-%Y")
            end   = datetime.strptime(end_str, "%d-%m-%Y")
        except ValueError:
            await message.reply(PROMPT)
            return

        if end <= start:
            await message.reply(PROMPT)
            return

        days = [start + timedelta(days=i) for i in range((end - start).days + 1)]

    else:
        # unmatched format
        await message.reply(PROMPT)
        return

    excel_buffer = BytesIO()
    
    pd.DataFrame().to_excel(excel_buffer , index = False)
    
    excel_buffer.seek(0)
    wb = openpyxl.load_workbook(excel_buffer)
#         wb.create_sheet('Sheet1')
    wb.create_sheet('Sheet2')
    ws = wb['Sheet1']
    ws.merge_cells('A1:A2')
    ws['A1'] = 'Ismlar'
    ws.column_dimensions['A'].width = 30
    ws.cell(row = 1 , column = 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws2 = wb['Sheet2']
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20
    
#         Generate images and names 

    for idx, name in enumerate(USERS_MAPPING.keys(), start=3):  # start=3 to start writing from A3
        cell = f'A{idx}'  # Construct cell reference
        ws[cell] = name     
        ws[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    images = USERS_MAPPING
    
    
    for i , day in enumerate(days):
        
        start = datetime.combine(day, time(4,0,0))
#             print(day , type(day))
        df1 , df2 = query_data_to_dataframe(day)
        cell1 , cell2, cell3, cell4 = xl_col_to_name(1 + i * 4), xl_col_to_name(2 + i * 4), xl_col_to_name(3 + i * 4), xl_col_to_name(4 + i * 4)
        
        ws.column_dimensions[cell1].width = 20
        ws.column_dimensions[cell2].width = 20
        ws.column_dimensions[cell3].width = 10
        ws.column_dimensions[cell4].width = 25
        
        ws.merge_cells(f'{cell1}1:{cell4}1' , day) 
        
        ws.cell(row = 1 , column = 2+i*4).value = day.strftime("%d-%m-%Y")
        ws.cell(row = 1 , column = 2+i*4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        ws.cell(row = 2 , column = 2+i*4).value = 'Kelgan vaqt'
        ws.cell(row = 2 , column = 2+i*4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        ws.cell(row = 2 , column = 3+i*4).value = 'Ketgan vaqt'
        ws.cell(row = 2 , column = 3+i*4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws.cell(row = 2 , column = 4+i*4).value = 'BALL'
        ws.cell(row = 2 , column = 4+i*4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws.cell(row = 2 , column = 5+i*4).value = 'Sabab'
        ws.cell(row = 2 , column = 5+i*4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        for j , name in enumerate(USERS_MAPPING.keys()):
            row = df1[df1['name'] == name]
            if not row.empty:
                images[name] = row['image'].iloc[0]
            
            if day.weekday() < 6:
                if not row.empty:
                    ws.cell(row = j+3, column = 2 + i * 4).value =  datetime.strptime(row["arrival_time"].iloc[0], '%Y-%m-%d %H:%M:%S').strftime("%H:%M:%S")
                    ws.cell(row = j+3, column = 2 + i * 4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

                    ws.cell(row = j+3, column = 3 + i * 4).value =  datetime.strptime(row["departure_time"].iloc[0], '%Y-%m-%d %H:%M:%S').strftime("%H:%M:%S")
                    ws.cell(row = j+3, column = 3 + i * 4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    
                    BALL = get_ball(start,
                                    datetime.strptime(row["arrival_time"].iloc[0] , "%Y-%m-%d %H:%M:%S"),                                
                                    datetime.strptime(row["departure_time"].iloc[0] , "%Y-%m-%d %H:%M:%S"))
                    
                    ws.cell(row = j+3, column = 4 + i * 4).value = BALL
                    ws.cell(row = j+3, column = 4 + i * 4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    try:
                        ws.cell(row = j+3, column = 5 + i * 4).value = row["reason"].iloc[0]
                    except:pass
                    ws.cell(row = j+3, column = 5 + i * 4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    
                    if datetime.strptime(row["arrival_time"].iloc[0] , "%Y-%m-%d %H:%M:%S") < start + timedelta(hours=6, minutes=30):
                        ws.cell(row = j+3, column = 2 + i * 4).fill = green
                    else:
                        ws.cell(row = j+3, column = 2 + i * 4).fill = orange
                    if datetime.strptime(row["departure_time"].iloc[0] , "%Y-%m-%d %H:%M:%S") > start + timedelta(hours=13, minutes=59) or BALL == "0":
                        ws.cell(row = j+3, column = 3 + i * 4).fill = green
                    else:
                        ws.cell(row = j+3, column = 3 + i * 4).fill = orange
                else:
                    ws.cell(row = j+3, column = 4 + i * 4).value =  str(-0.5)
                    ws.cell(row = j+3, column = 4 + i * 4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    
                    ws.cell(row = j+3, column = 2 + i * 4).fill = red
                    ws.cell(row = j+3, column = 3 + i * 4).fill = red
                    
                if day.weekday() == 5:
                    border = openpyxl.styles.Border(left = openpyxl.styles.Side(style='thick'),
                                                            right = openpyxl.styles.Side(style='thick'),
                                                            top = openpyxl.styles.Side(style='thick'),
                                                            bottom = openpyxl.styles.Side(style='thick'))
                    
                    ws.cell(row = j+3, column = 2 + i * 4).border = border
                    ws.cell(row = j+3, column = 3 + i * 4).border = border
                    ws.cell(row = j+3, column = 4 + i * 4).border = border
                    ws.cell(row = j+3, column = 5 + i * 4).border = border
                    


#                         if datetime.strptime(row["departure_time"].iloc[0] , "%Y-%m-%d %H:%M:%S") > start + timedelta(hours=13, minutes=59):
#                             ws.cell(row = j+3, column = 3 + i * 2).fill = green
#                         elif (datetime.strptime(row["arrival_time"].iloc[0] , "%Y-%m-%d %H:%M:%S") < start + timedelta(hours=5)) and (datetime.strptime(row["departure_time"].iloc[0] , "%Y-%m-%d %H:%M:%S") > datetime.strptime(row["arrival_time"].iloc[0] , "%Y-%m-%d %H:%M:%S") + timedelta(hours = 7 , minutes = 59)):
#                             ws.cell(row = j+3, column = 3 + i * 2).fill = green
#                         else:
#                             ws.cell(row = j+3, column = 3 + i * 2).fill = orange
#                     else:
#                         ws.cell(row = j+3, column = 2 + i * 2).fill = red
#                         ws.cell(row = j+3, column = 3 + i * 2).fill = red
                
#                     for weekends
            else:
                if not row.empty:
                    
                    ws.cell(row = j+3, column = 2 + i * 4).value =  datetime.strptime(row["arrival_time"].iloc[0], '%Y-%m-%d %H:%M:%S').strftime("%H:%M:%S")
                    ws.cell(row = j+3, column = 3 + i * 4).value =  datetime.strptime(row["departure_time"].iloc[0], '%Y-%m-%d %H:%M:%S').strftime("%H:%M:%S")
                    
                    ws.cell(row = j+3, column = 2 + i * 4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    ws.cell(row = j+3, column = 3 + i * 4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
#             add unknown in  worksheet2
        if len(df2)>0:
            next_row = ws2.max_row + 1 if ws2.max_row > 1 else 1
            
            ws2.merge_cells(f'A{next_row}:D{next_row}' , day)
            
            ws2.cell(row = next_row , column = 1).value = day.strftime("%d-%m-%Y")
            ws2.cell(row = next_row , column = 1).fill = green
            ws2.cell(row = next_row , column = 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            for row in df2.values.tolist():
                ws2.append(row[:-1])

                if pd.notna(row[3]) and os.path.isfile(row[3]):

                    img = Image(row[3])
                    aspect_ratio = img.width / img.height
                    img.height = 50  # Set height
                    img.width = aspect_ratio * img.height
                    # Adjust image size if needed
                    cell_ref = ws2.cell(row=ws2.max_row, column=4)
                    ws2.add_image(img, cell_ref.coordinate)          
                    ws2.row_dimensions[ws2.max_row].height = 50
        
#         add images worksheet1
    for idx, name in enumerate(images.keys(), start=3):
        if isinstance(images[name], str) and os.path.isfile(images[name]):
            img = Image(images[name])
            aspect_ratio = img.width / img.height
            img.height = 50  # Set height
            img.width = aspect_ratio * img.height
            ws.add_image(img , f'A{idx}')
            ws.row_dimensions[idx].height = 50   
            
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    discord_file = discord.File(fp=excel_buffer, filename=f"{text}.xlsx")

    # send it to the same channel
    await message.channel.send(file=discord_file)


if __name__ == '__main__':
    bot.run('')

                
            
            














